#!/usr/bin/env python3
from clockmmu import ClockMMU
from lrummu import LruMMU
from randmmu import RandMMU
import sys
import openpyxl  # built-in module for Excel operations
#from itertools import islice


def main():
    PAGE_OFFSET = 12  # page is 2^12 = 4KB

    ############################
    # Check input parameters   #
    ############################

    if len(sys.argv) < 2:
        print("Usage: python memsim.py settings_file")
        return

    settings_file = sys.argv[1]
    wb = openpyxl.load_workbook(settings_file)
    sheet = wb.active

    # Create an output workbook and set column titles
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    titles = ["Trace File Type", "Replacement Mode", "total memory frames", "events in trace", 
              "total disk reads", "total disk writes", "page fault rate", "cache hit rate"]
    ws_out.append(titles)

#    for row in islice(sheet.iter_rows(values_only=True), 1, None):
    for row in sheet.iter_rows(values_only=True):
        trace_file_type, frames, replacement_mode, debug_mode = row
        
#        if frames < 1:
#            ws_out.append(["Error: Frame number must be at least 1"])
#            continue

        # Setup MMU based on replacement mode
        if replacement_mode == "rand":
            mmu = RandMMU(frames)
        elif replacement_mode == "lru":
            mmu = LruMMU(frames)
        elif replacement_mode == "clock":
            mmu = ClockMMU(frames)
        else:
            continue  # Skip this row and move to the next if replacement mode is invalid

        # Set debug mode
        if debug_mode == "debug":
            mmu.set_debug()
        elif debug_mode == "quiet":
            mmu.reset_debug()
        else:
            continue  # Skip this row and move to the next if debug mode is invalid

        no_events = 0
        with open(trace_file_type, 'r') as trace_file:
            for trace_line in trace_file:
                trace_cmd = trace_line.strip().split(" ")
                logical_address = int(trace_cmd[0], 16)
                page_number = logical_address >> PAGE_OFFSET

                # Process read or write
                if trace_cmd[1] == "R":
                    mmu.read_memory(page_number)
                elif trace_cmd[1] == "W":
                    mmu.write_memory(page_number)
                else:
                    continue  # Skip this line and move to the next if the trace command is invalid

                no_events += 1

        # Calculate metrics
        page_fault_rate = mmu.get_total_page_faults() / no_events
        cache_hit_rate = 1 - page_fault_rate  # Assuming cache hit rate = 1 - page fault rate. Adjust as necessary.

        # Append results to Excel output
        results = [trace_file_type, replacement_mode, frames, no_events, mmu.get_total_disk_reads(), mmu.get_total_disk_writes(), page_fault_rate, cache_hit_rate]
        ws_out.append(results)

    wb_out.save('results.xlsx')  # save to results.xlsx

if __name__ == "__main__":
    main()
